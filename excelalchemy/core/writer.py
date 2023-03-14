"""负责将 pandas 写入 Excel 文件"""
import base64
from collections import defaultdict
from math import ceil
from tempfile import NamedTemporaryFile
from typing import BinaryIO
from typing import cast

from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from pandas import ExcelWriter

from excelalchemy.const import BACKGROUND_ERROR_COLOR
from excelalchemy.const import BACKGROUND_REQUIRED_COLOR
from excelalchemy.const import CHARACTER_WIDTH
from excelalchemy.const import DEFAULT_SHEET_NAME
from excelalchemy.const import FONT_READ_COLOR
from excelalchemy.const import HEADER_HINT
from excelalchemy.const import REASON_COLUMN_LABEL
from excelalchemy.const import RESULT_COLUMN_LABEL
from excelalchemy.exc import ExcelCellError
from excelalchemy.types.field import FieldMetaInfo
from excelalchemy.types.identity import Base64Str
from excelalchemy.types.identity import ColumnIndex
from excelalchemy.types.identity import Label
from excelalchemy.types.identity import RowIndex
from excelalchemy.types.identity import UniqueLabel
from excelalchemy.types.result import ValidateRowResult
from excelalchemy.types.value import EXCEL_CHOICE_VALUE_TYPE
from excelalchemy.util.file import add_excel_prefix
from excelalchemy.util.file import value_is_nan

# pandas 认为 Excel 的第一行是 0, 第一列是 0
PANDAS_EXCEL_INDEX_START_AT = 0
PANDAS_WRITE_START_AT = PANDAS_EXCEL_INDEX_START_AT + 1  # 从第二行开始写入数据,第一行写入 HEADER_HINT

# openpyxl 认为 Excel 的第一行是 1, 第一列是 1
OPENPYXL_EXCEL_INDEX_START_AT = 1  # Excel 从 1 开始

# 写入 HEADER_HINT 的 位置，基于 openpyxl 的索引
HEADER_HINT_ROW_INDEX = 1
HEADER_HINT_COL_INDEX = 1
HEADER_HINT_LINE_COUNT = 1  # HEADER_HINT 占用的行数

# 最多只能设置 16384 行数据的选项
MAX_OPTION_ROW_COUNT = 16384

# 简单表头的行数
SIMPLE_HEADER_ROW_COUNT = 1

# 合并表头的行数
MERGE_HEADER_ROW_COUNT = 2

# row_write_offset : 写入 Excel 时，从第几行开始写入
# column_write_offset : 写入 Excel 时，从第几列开始写入


def _get_file(file: BinaryIO | None = None) -> BinaryIO:
    """生成临时文件"""
    return cast(BinaryIO, file or NamedTemporaryFile())


# pylint: disable=too-many-locals
def _write_simple_header(
    df: DataFrame,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    file: BinaryIO,
    sheet_name: str,
    column_write_offset: int = 0,
    row_write_offset: int = 0,
    close_file: bool = True,
    writer: ExcelWriter | None = None,
    option_start_at: int = OPENPYXL_EXCEL_INDEX_START_AT + HEADER_HINT_LINE_COUNT + SIMPLE_HEADER_ROW_COUNT,
) -> BinaryIO:
    """写入简单的表头(没有合并的表头)"""

    writer = writer or ExcelWriter(file, engine='openpyxl')
    # pyright: reportUnknownMemberType=false
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=PANDAS_WRITE_START_AT)
    worksheet: Worksheet = writer.sheets[sheet_name]

    for openpyxl_col_index, column in enumerate(
        df.columns[column_write_offset:],
        start=column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT,
    ):  # pyright: reportUnknownArgumentType=false
        field_meta = field_meta_mapping[column]
        comment_text = field_meta.value_type.comment(field_meta)
        comment = Comment(
            text=comment_text,
            author='Teletraan',
            height=sum(ceil(len(line) / 20) for line in comment_text.splitlines()) * 28,
            width=300,
        )
        cell = worksheet.cell(row=row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT, column=openpyxl_col_index)
        if comment_text:
            cell.comment = comment
        if field_meta.required:
            cell.fill = PatternFill(start_color=BACKGROUND_REQUIRED_COLOR, fill_type='solid')  # 如果是必填项，设置背景颜色

        cell.font = Font(bold=True)  # 字体加粗
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.number_format = numbers.FORMAT_TEXT
        # 设置列为文本格式
        worksheet.column_dimensions[get_column_letter(openpyxl_col_index)].number_format = numbers.FORMAT_TEXT

        # 设置列的下拉值可选项, 只支持单选
        if field_meta.options and field_meta.value_type in EXCEL_CHOICE_VALUE_TYPE:
            column_letter = get_column_letter(openpyxl_col_index)
            data_validation = DataValidation(
                type='list',
                formula1=f'"{",".join(x.name for x in field_meta.options)}"',
                allow_blank=not field_meta.required,
                # option_start_at + 1 表头行不需要下拉选项
                sqref=f'{column_letter}{option_start_at + 1}:{column_letter}{MAX_OPTION_ROW_COUNT}',
                error='请从下拉列表中选择',
                errorTitle=f'【{field_meta.label}】列填写错误',
            )
            worksheet.add_data_validation(data_validation)

    if close_file:
        writer.close()

    return file


def _write_comment_header(
    df: DataFrame,
    file: BinaryIO,
    sheet_name: str,
    close_file: bool = True,
    writer: ExcelWriter | None = None,
) -> BinaryIO:
    """写入 HEADER_HINT"""

    writer = writer or ExcelWriter(file, engine='openpyxl')
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=PANDAS_WRITE_START_AT)
    worksheet: Worksheet = writer.sheets[sheet_name]
    cell = worksheet.cell(row=HEADER_HINT_ROW_INDEX, column=HEADER_HINT_COL_INDEX)
    cell.value = HEADER_HINT
    cell.font = Font(size=16)
    cell.alignment = Alignment(wrap_text=True)
    worksheet.merge_cells(
        start_row=HEADER_HINT_ROW_INDEX,
        start_column=HEADER_HINT_COL_INDEX,
        end_row=HEADER_HINT_ROW_INDEX,
        end_column=len(df.columns),
    )
    # pyright: reportGeneralTypeIssues=false
    worksheet.row_dimensions[HEADER_HINT_ROW_INDEX].height = 120

    if close_file:
        writer.close()

    return file


def _write_merged_header(  # pragma: no mccabe
    df: DataFrame,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    file: BinaryIO,
    sheet_name: str,
    column_write_offset: int = 0,
    row_write_offset: int = 0,
    close_file: bool = True,
    writer: ExcelWriter | None = None,
) -> BinaryIO:
    """写入含有合并的表头"""

    writer = writer or ExcelWriter(file, engine='openpyxl')
    worksheet: Worksheet = writer.sheets[sheet_name]

    # 写入注释需要在合并表头之前
    _write_simple_header(
        df,
        field_meta_mapping,
        file,
        sheet_name,
        column_write_offset,
        row_write_offset,
        close_file=False,
        writer=writer,
        option_start_at=OPENPYXL_EXCEL_INDEX_START_AT + HEADER_HINT_LINE_COUNT + MERGE_HEADER_ROW_COUNT,
    )
    # 第一遍遍历，找出纵向合并的单元格
    start_row = row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
    for openpyxl_col_index, column in enumerate(
        df.columns[column_write_offset:],
        start=column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT,
    ):
        field_meta = field_meta_mapping[column]
        if field_meta.label == field_meta.parent_label:
            # 如果 label 和 parent_label 相同，说明需要上下合并
            worksheet.merge_cells(
                start_row=start_row,
                start_column=openpyxl_col_index + column_write_offset,
                end_row=start_row + 1,  # +1 表示合并两行
                end_column=openpyxl_col_index + column_write_offset,
            )
        worksheet.cell(
            row=start_row,
            column=openpyxl_col_index + column_write_offset,
        ).number_format = numbers.FORMAT_TEXT

    # 第二遍遍历，找出横向合并的单元格
    counter: dict[Label, int] = defaultdict(int)
    for field_meta in field_meta_mapping.values():
        if field_meta.parent_label is None:
            raise RuntimeError('运行时 parent_label 不能为空')
        counter[field_meta.parent_label] += 1

    for openpyxl_col_index, column in enumerate(
        df.columns[column_write_offset:],
        start=column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT,
    ):
        field_meta = field_meta_mapping[column]
        if field_meta.parent_label is None:
            raise RuntimeError('运行时 parent_label 不能为空')
        if field_meta.label != field_meta.parent_label and field_meta.offset == 0:
            # 如果 label 和 parent_label 不同，说明需要左右合并
            # 首先设置值
            cell = worksheet.cell(row=start_row, column=openpyxl_col_index + column_write_offset)
            cell.value = field_meta.parent_label
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 然后合并单元格
            worksheet.merge_cells(
                start_row=start_row,
                start_column=openpyxl_col_index + column_write_offset,
                end_row=start_row,
                end_column=openpyxl_col_index + column_write_offset + counter[field_meta.parent_label] - 1,
            )

    if close_file:
        writer.close()

    return file


def _get_parsed_value(
    df: DataFrame,
    row_index: int,
    col_index: int,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
) -> str:
    """用于把 pandas 读取的 Excel 之后的数据，转回用户可识别的数据"""

    cell_value = df.iloc[row_index, col_index]

    if cell_value is None or value_is_nan(cell_value):
        return ''  # parse None for end-user
    col_label = cast(UniqueLabel, df.columns[col_index])
    if col_label not in field_meta_mapping:
        return str(cell_value)
    field_meta = field_meta_mapping[col_label]
    cell_value = field_meta.value_type.deserialize(cell_value, field_meta)

    return str(cell_value)


# pylint: disable=too-many-locals
def _write_value_mark_error(  # pragma: no mccabe
    df: DataFrame,
    errors: dict[RowIndex, dict[ColumnIndex, list[ExcelCellError]]],
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    file: BinaryIO,
    sheet_name: str,
    row_write_offset: int = 0,
    column_write_offset: int = 0,
    close_file: bool = True,
    writer: ExcelWriter | None = None,
    pands_data_start_index: int = 0,
) -> BinaryIO:
    """写入错误标记，并把对应位置标红"""

    writer = writer or ExcelWriter(file, engine='openpyxl')
    worksheet: Worksheet = writer.sheets[sheet_name]

    for row_index, cols in errors.items():
        for col_index, exceptions in cols.items():
            if not exceptions:
                continue

            openpyxl_col_index = col_index + column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
            openpyxl_row_index = row_index + row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT

            # 设置单元格背景为红色
            cell = worksheet.cell(row=openpyxl_row_index, column=openpyxl_col_index)
            cell.fill = PatternFill(
                start_color=BACKGROUND_ERROR_COLOR,
                end_color=BACKGROUND_ERROR_COLOR,
                fill_type='solid',
            )
            cell.alignment = Alignment(wrap_text=True)

    col_width_mapping: dict[ColumnIndex, float] = defaultdict(float)
    for row_index_ in range(pands_data_start_index, df.shape[0]):  # iterate over rows
        for column_index_ in range(df.shape[1]):  # iterate over columns
            openpyxl_col_index = column_index_ + column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
            openpyxl_row_index = row_index_ + row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT

            cell = worksheet.cell(row=openpyxl_row_index, column=openpyxl_col_index)
            cell.value = _get_parsed_value(df, row_index_, column_index_, field_meta_mapping)
            cell.number_format = numbers.FORMAT_TEXT
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if RESULT_COLUMN_LABEL == df.columns[column_index_] and cell.value == str(ValidateRowResult.FAIL):
                cell.font = Font(color=FONT_READ_COLOR)  # 设置文字颜色为红色
            if REASON_COLUMN_LABEL == df.columns[column_index_]:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            col_width_mapping[ColumnIndex(openpyxl_col_index)] = max(
                col_width_mapping[ColumnIndex(openpyxl_col_index)],
                max(len(str(x)) for x in str(cell.value).split('\n')),
                len(str(df.columns[column_index_])),
            )

    for openpyxl_col_index, width in col_width_mapping.items():
        worksheet.column_dimensions[get_column_letter(openpyxl_col_index)].width = round(
            (width + 4) * CHARACTER_WIDTH, 2
        )

    if close_file:
        writer.close()
    return file


def render_simple_header_excel(
    df: DataFrame,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    sheet_name: str = DEFAULT_SHEET_NAME,
    file: BinaryIO | None = None,
    close_file: bool = True,
    column_write_offset: int = 0,
) -> str:
    """把表头写入 Excel 文件"""
    if file is None:
        close_file = True

    tmp = _get_file(file)
    writer = ExcelWriter(tmp, engine='openpyxl')
    _write_comment_header(df, tmp, sheet_name, writer=writer, close_file=False)
    _write_simple_header(
        df,
        field_meta_mapping,
        tmp,
        sheet_name,
        column_write_offset,
        row_write_offset=HEADER_HINT_LINE_COUNT,
        writer=writer,
        close_file=False,
    )

    writer.close()
    tmp.seek(0)
    content = base64.b64encode(tmp.read()).decode()
    if close_file:
        tmp.close()
    return add_excel_prefix(content)


def render_merged_header_excel(
    df: DataFrame,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    sheet_name: str = DEFAULT_SHEET_NAME,
    file: BinaryIO | None = None,
    close_file: bool = True,
    column_write_offset: int = 0,
) -> str:
    """把合并的表头写入 Excel 文件"""
    if file is None:
        close_file = True

    tmp = _get_file(file)
    writer = ExcelWriter(tmp, engine='openpyxl')
    _write_comment_header(df, tmp, sheet_name, writer=writer, close_file=False)
    _write_merged_header(
        df,
        field_meta_mapping,
        tmp,
        sheet_name,
        column_write_offset,
        row_write_offset=HEADER_HINT_LINE_COUNT,
        writer=writer,
        close_file=False,
    )

    writer.close()  # writer 需要先 close，否则无法读取到数据
    tmp.seek(0)
    content = base64.b64encode(tmp.read()).decode()

    if close_file:
        tmp.close()
    return add_excel_prefix(content)


def render_data_excel(
    df: DataFrame,
    errors: dict[RowIndex, dict[ColumnIndex, list[ExcelCellError]]],
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    sheet_name: str = DEFAULT_SHEET_NAME,
    file: BinaryIO | None = None,
    close_file: bool = True,
    has_merged_header: bool = False,
) -> Base64Str:
    if file is None:
        close_file = True

    tmp = _get_file(file)
    writer = ExcelWriter(tmp, engine='openpyxl')

    _write_comment_header(df, tmp, sheet_name, writer=writer, close_file=False)
    if has_merged_header:
        pands_data_start_index = 1
        _write_merged_header(
            df,
            field_meta_mapping,
            tmp,
            sheet_name,
            row_write_offset=HEADER_HINT_LINE_COUNT,
            writer=writer,
            close_file=False,
        )
    else:
        pands_data_start_index = 0
        _write_simple_header(
            df,
            field_meta_mapping,
            tmp,
            sheet_name,
            row_write_offset=HEADER_HINT_LINE_COUNT,
            writer=writer,
            close_file=False,
        )
    _write_value_mark_error(
        df,
        errors,
        field_meta_mapping,
        tmp,
        sheet_name,
        row_write_offset=HEADER_HINT_LINE_COUNT + 1,  # 表头 1 行，HEADER_HINT 一行
        writer=writer,
        close_file=False,
        pands_data_start_index=pands_data_start_index,
    )

    writer.close()
    tmp.seek(0)
    content = base64.b64encode(tmp.read()).decode()
    if close_file:
        tmp.close()
    return Base64Str(add_excel_prefix(content))
